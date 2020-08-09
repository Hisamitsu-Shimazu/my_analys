import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# 自作モジュール
import module.excel_style as excel_style

# styleの設定
def set_style(cell, s):
    '''
    wsのcellに対してstyleの設定を行う。
    
    Paramaters
    ----------
    cell : openpyxl.ws.cell
        設定対象のセル
    s : str
        styleを表す文字列
        styleに関してはexcel_style.pyに記載
    
    '''
    # スタイル設定
    cell.font = excel_style.style[s][0]  # Font
    cell.fill = excel_style.style[s][1]   # PatternFill
    cell.border = excel_style.style[s][2]  # Border
    cell.alignment = excel_style.style[s][3]  # Alignment
    
# 値とstyleを設定
def set_cell(ws, cell:tuple, value=None, style='normal'):
    '''
    値とスタイルを設定する。
    基本的にセルの編集はこれを用いる。
    
    Paramaters
    ----------
    ws : openpyxl.ws
        編集対象のws
    cell : tuple
        セルのindexをタプルで指定
        (行, 列)
    value : all
        セルに代入する値
    style : str
        設定するスタイル
        詳細はexcel_style.pyを参照
        
    '''
    ws.cell(cell[0], cell[1]).value = value # 値
    set_style(ws.cell(cell[0], cell[1]), style) # style
    
# データフレームの挿入
def put_dataframe(ws, df, start:tuple, mode='normal'):
    '''
    エクセルシートにデータフレームを挿入する。
    
    Paramaters
    ----------
    ws : openpyxl.ws
        編集対象のws
    df : pd.DataFrame
        挿入するdataframe
    start : tuple
        挿入を開始するセル
    mode : str
        nomal : 通常モード、欠損をnanで埋める。
        describe : 統計量モード、欠損を-で埋める。
    
    '''
    # 枠
    ## 左上のセル
    set_cell(ws, cell=(start[0], start[1]), value='', style='style_header')
    ## カラム代入
    for i,c in enumerate(df.columns):
        set_cell(ws, cell=(start[0], start[1]+1+i), value=str(c), style='style_header')
    ## インデックス代入
    for i,r in enumerate(df.index):
        set_cell(ws, cell=(start[0]+1+i, start[1]), value=str(r), style='style_header')
        
    # 値代入
    ## 行で回す
    for i in range(df.shape[0]):
        ## 列で回す
        for j in range(df.shape[1]):
            ## 欠損の処理
            if df.isnull().iat[i,j]:
                if mode == 'describe':
                    set_cell(ws, cell=(start[0]+1+i, start[1]+1+j), value='-', style='style_grid')
                else :
                    set_cell(ws, cell=(start[0]+1+i, start[1]+1+j), value='Nan', style='style_grid')
            ## 欠損以外
            else :
                set_cell(ws, cell=(start[0]+1+i, start[1]+1+j), value=df.iat[i,j], style='style_grid')

# 幅調整
def adjust_width(ws):
    '''
    自動でシートの幅調整を行う関数
    
    Paramaters
    ----------
    ws : openpxl
        編集対象のws
        
    '''
    # 列で回す
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        ## 値が入っているセルをすべて回す
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ## 調整幅を設定
        ## 最大幅の1.2倍
        adjusted_width = (max_length + 2) * 1.2
        
        ## 幅が広くなりすぎないように上限を35に設定
        if adjusted_width > 35:
            ws.column_dimensions[get_column_letter(column)].width = 35
        else:
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

# 列番号、行番号から文字列取得
def get_letter(r, c):
    '''
    列番号、行番号からセルを表す文字列を取得
    
    Paramaters
    ----------
    r : int
        行
    c : int
        列
    
    '''
    return openpyxl.utils.get_column_letter(c)+str(r)
            
# 画像貼り付け
def put_img(ws, fig, c, start:tuple, img_dir):
    '''
    画像を張り付ける関数
    
    Paramaters
    ----------
    ws : openpyxl.ws
        編集対象のws
    fig : pd.figure
        貼り付け用fig
    c : str
        カラム名
    start : tuple
        挿入を開始するセル
    img_dir : pathlib.Path
        画像保存先のパス
        
    '''
    # 一旦画像保存
    fig.savefig(img_dir/f'{c}_fig.png', transparent=True, bbox_inches='tight')
    # 画像保存
    img = openpyxl.drawing.image.Image(img_dir/f'{c}_fig.png')
    # 貼り付け
    ws.add_image(img, get_letter(start[0], start[1]))
    