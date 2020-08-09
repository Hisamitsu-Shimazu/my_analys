'''
openpyxlを用いてstyleを設定する。
edit_excel.pyのset_style()関数で使用する。

・Font
・PatternFill
・Border
・Alignment

に関して調整を行う。
'''

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

style = {
    # 普通
    'normal' : [
        Font(
        ),
        PatternFill(
        ),
        Border(
        ),
        Alignment(
        )
    ],
    
    # タイトル
    'style_title' : [
        Font(
            bold = True,
            underline='double'
        ),
        PatternFill(
            
        ),
        Border(
            
        ),
        Alignment(
            vertical = 'center'
        )
    ],
    
    # ヘッダー
    'style_header' : [
        Font(
            bold = True,
            underline='none',
            color = '000000'
        ),
        PatternFill(
            'solid',
            fgColor='badcad'
        ),
        Border(
            left=Side(border_style='thin',
                       color='000000'),
            right=Side(border_style='thin',
                       color='000000'),
            top=Side(border_style='thin',
                       color='000000'),
            bottom=Side(border_style='thin',
                       color='000000')
        ),
        Alignment(
            horizontal = 'center', 
            vertical = 'center',
            wrap_text = False
        )
    ],
    
    # サマリ
    'style_summary' : [
        Font(
            bold = False,
            underline='single',
            color = '1e90ff'
        ),
        PatternFill(
            'solid',
            fgColor='ffffff'
        ),
        Border(
            left=Side(border_style='thin',
                       color='000000'),
            right=Side(border_style='thin',
                       color='000000'),
            top=Side(border_style='thin',
                       color='000000'),
            bottom=Side(border_style='thin',
                       color='000000')
        ),
        Alignment(
            horizontal = 'center', 
            vertical = 'center',
            wrap_text = False
        )
    ],
    
    # 表
    'style_grid' : [
        Font(
            bold = False,
            underline='none',
            color = '000000'
        ),
        PatternFill(
            'solid',
            fgColor='ffffff'
        ),
        Border(
            left=Side(border_style='thin',
                       color='000000'),
            right=Side(border_style='thin',
                       color='000000'),
            top=Side(border_style='thin',
                       color='000000'),
            bottom=Side(border_style='thin',
                       color='000000')
        ),
        Alignment(
            horizontal = 'center', 
            vertical = 'center',
            wrap_text = False
        )
    ],
    
    # 画像
    'style_img' : [
        Font(
            bold = False,
            underline='none',
            color = '000000'
        ),
        PatternFill(
            'solid',
            fgColor='ffffff'
        ),
        Border(
            left=Side(border_style='thin',
                       color='000000'),
            right=Side(border_style='thin',
                       color='000000'),
            top=Side(border_style='thin',
                       color='000000'),
            bottom=Side(border_style='thin',
                       color='000000')
        ),
        Alignment(
            horizontal = 'center', 
            vertical = 'center',
            wrap_text = False
        )
    ]
}