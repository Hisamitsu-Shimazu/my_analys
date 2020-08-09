# パッケージ読み込み
import os
import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib
import openpyxl
from openpyxl.formatting.rule import DataBarRule
from pathlib import Path
import warnings
warnings.simplefilter('ignore')

# 自作モジュールの読み込み
sys.path.append('./module')
import edit_excel as ee

# 各種設定
input_dir = Path('../data/')
img_dir = Path('../output/img/')
output_dir = Path('../output/')

def main():
    ## 実行部
    # 対象ファイル
    file_list = os.listdir(input_dir)

    for file_name in file_list:
        if file_name=='.gitkeep.txt' : continue
        
        print(f'----- {file_name} -----')
        # データ読み込み
        df = read_data(file_name)
        if df is False :
            print(f'# {file_name} is not csv or pickle.')
            continue
            
        print(f'# {file_name} has read.')

        # ワークブック作成
        wb = create_workbook(file_name)

        # 最初のページ
        create_first_sheet(wb, file_name, df)
        print(f'# create samaly sheet.')

        # 次のページ
        create_second_sheet(wb, df)
        print(f'# create describe sheet.')
        
        # データ例ページ
        create_data_show_sheet(wb, df)
        print(f'# create data-show sheet.')
        
        # カラムごとのシート
        for i, c in enumerate(df.columns):
            
            if i == 0: print(f'# column : {c}...', end='')
            else : print(f'\r# column : {c}...', end='')
            
            create_column_sheet(wb, df[c])
        
        print('\r# column : completed !')
        # ワークブックの保存
        wb.save(output_dir/(file_name.split('.')[0]+'.xlsx'))
        print('# {} has saved.'.format(file_name.split('.')[0]+'.xlsx'))
        
        
def read_data(file_name : str):
    '''
    データの読み込み
    
    Paramaters
    ----------
    file_name : str
        分析対象のファイル名
        
    input_dir : Path
        データ格納先までのパス
    
    Return
    ------
    df : pd.DataFrame | False
    
    '''
    
    if file_name.split('.')[1]=='csv':
        df = pd.read_csv(input_dir/file_name, engine='python', index_col=0) # csv
    elif file_name.split('.')[1]=='pickle':
        df = pd.read_pickle(input_dir/file_name) # pickle
    else:
        df = False
    return df

def create_workbook(file_name:str):
    '''
    ワークブック作成
    
    Paramaters
    ----------
    file_name : str
    
    Return
    ------
    wb : openpyxl.workbook
    
    '''
    
    # ワークブックの作成
    wb = openpyxl.Workbook()
    f_name = file_name.split('.')[0]+'.xlsx'
    wb.save(output_dir/f_name)

    # ワークブックの呼び出し
    wb = openpyxl.load_workbook(output_dir/f_name)
    return wb
    
def create_first_sheet(wb, file_name:str, df:pd.DataFrame):
    '''
    サマリシート作成
    
    Paramaters
    ----------
    ws : openpyxl.workbook
        ワークブック
        
    file_name : str
        分析対象のファイル名
    
    '''
    # 最初のページ
    ws = wb['Sheet']
    ws.sheet_properties.tabColor = 'c0c0c0'
    
    # シート名の変更
    ws.title = 'サマリ'
    
    # タイトル
    ee.set_cell(ws, cell=(2,2), value='サマリ', style='style_title')
    
    # 上部情報追加
    ee.set_cell(ws, cell=(3,2), value='ファイル名', style='style_header')
    ee.set_cell(ws, cell=(3,3), value=file_name, style='style_grid')

    # カラム情報
    ## ヘッダー
    for i, c in enumerate(['No.', '論理名', '物理名', '解釈', '備考']):
        ee.set_cell(ws, cell=(4, 2+i), value=c, style='style_header')
    
    f_name = file_name.split('.')[0]+'.xlsx'
    ## インデックス
    for i, c in enumerate(df.columns):
        ee.set_cell(ws, cell=(5+i, 2), value=i, style='style_header')
        ee.set_cell(ws, cell=(5+i, 3), value=c, style='style_summary')
        ee.set_cell(ws, cell=(5+i, 4), style='style_grid')
        ee.set_cell(ws, cell=(5+i, 5), style='style_grid')
        ee.set_cell(ws, cell=(5+i, 6), style='style_grid')
        # ハイパーリンク
        ws[ee.get_letter(5+i, 3)].hyperlink = f'{f_name}#{c}!A1'
        
    # ウィンドウ枠の固定
    ws.freeze_panes = 'C5'
        
    # 幅、高さ調整
    ws.row_dimensions[1].height = 5
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
def create_second_sheet(wb, df:pd.DataFrame):
    '''
    2番目のシート作成
    
    Paramaters
    ----------
    wb : openpyxl.workbook 
        ワークブック
        
    '''
    
    # 次のページ
    ws = wb.create_sheet(title='統計量')
    ws.sheet_properties.tabColor = '87cefa'
    
    ee.set_cell(ws, cell=(2, 2), value='統計量', style='style_title')
    
    tmp1 = pd.DataFrame(
        {
            '欠損' : df.isnull().sum().values,
            '欠損率(%)' : df.isnull().sum().values / len(df) * 100
        },
        index = df.columns
    )
    tmp2 = df.describe(include='all').T
    tmp2.columns = ['行数', 'ﾕﾆｰｸ数', '先頭行', '最頻値', '平均', '標準偏差', '最小', '25%', '中央値', '75%', '最大']
    tmp2 = tmp2.drop(['先頭行'], axis=1)
    
    tmp3 = pd.DataFrame(
        {
            'メモ' : [np.nan] * len(df.columns)
        },
        index = df.columns
    )
    
    # 統計量記入
    describe = pd.concat([tmp1, tmp2, tmp3], axis=1)
    for c in describe.columns:
        if c not in ['欠損', 'メモ']:
            describe[c] = describe[c].astype(float).round(3)
    ee.put_dataframe(ws, describe, start=(4, 2), mode='describe')
    
    # データバー
    ws.conditional_formatting.add(f'D5:D{5+len(df.columns)}',
                                   DataBarRule( start_type='num', start_value=0, end_type='num', end_value='100', 
                                       color="b94047", showValue="None", minLength=None, maxLength=None))

    # 行数、列数
    ee.set_cell(ws, cell=(3, 2), value='行', style='style_header')
    ee.set_cell(ws, cell=(3, 3), value=df.shape[0], style='style_grid')
    ee.set_cell(ws, cell=(3, 4), value='列', style='style_header')
    ee.set_cell(ws, cell=(3, 5), value=df.shape[1], style='style_grid')

    # 先頭、末尾
    #ee.put_dataframe(ws, df.head(20), start=(19,2))
    #ee.put_dataframe(ws, df.tail(20), start=(41,2))

    # ウィンドウ枠の固定
    ws.freeze_panes = 'C5'
    
    # 幅調整
    ee.adjust_width(ws)
    ws.row_dimensions[1].height = 5
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['O'].width = 30

def create_data_show_sheet(wb, df:pd.DataFrame):
    # データ例を見る
    ws = wb.create_sheet(title='データ例')
    ws.sheet_properties.tabColor = 'f0e68c'
    
    ee.set_cell(ws, cell=(2, 2), value='データ例', style='style_title')
    
    tmp1 = df.head(20)
    tmp2 = df.tail(20)
    ee.put_dataframe(ws, tmp1, start=(3,2))
    ee.put_dataframe(ws, tmp2, start=(25,2))
    
    # 高さ幅調整
    ee.adjust_width(ws)
    ws.freeze_panes = 'A4'
    ws.row_dimensions[1].height = 5
    ws.column_dimensions['A'].width = 1
    
def create_column_sheet(wb, d:pd.Series):
    '''
    カラムごとのシート作成
    
    Paramaters
    ----------
    wb : openpyxl.workbook
        ワークブック
    d : pd.Series
        対象のカラムのシリーズ
    
    '''
    # シート作成
    ws = wb.create_sheet(title=d.name)
    ws.sheet_properties.tabColor = 'f0e68c'
    
    ee.set_cell(ws, cell=(2, 2), value=d.name, style='style_title')
    
    # 共通処理
    ee.set_cell(ws, cell=(3, 2), value='論理名', style='style_header')
    ee.set_cell(ws, cell=(3, 3), style='style_grid')
    ee.set_cell(ws, cell=(3, 4), value='物理名', style='style_header')
    ee.set_cell(ws, cell=(3, 5), style='style_grid')
    ee.set_cell(ws, cell=(3, 6), value='欠損', style='style_header')
    
    # 欠損
    if d.isnull().sum() != 0:
        ee.set_cell(ws, cell=(3, 7), value=d.isnull().sum(), style='style_grid')
    else :
        ee.set_cell(ws, cell=(3, 7), value='なし', style='style_grid')
    
    # 可視化
    
    if d.dtype == object:
        if d.nunique()<=30:
            # 棒グラフ、帯グラフ
            fig = bar_plot(d)
        else:
            # ﾕﾆｰｸ数過多
            fig = too_much_nuniques(d)
        
    else :
        if d.nunique()<=30:
            # 棒グラフ
            fig = bar_plot(d)
        else:
            # ヒストグラム
            fig = hist_plot(d)
    
    ## fig貼り付け
    ee.put_img(ws, fig, d.name, (4,2), img_dir)
    ee.set_style(ws.cell(4,2), 'style_grid')
    ws.merge_cells('B4:G22')
    
    # 欠損率可視化
    fig = plt.figure(figsize=(3.5, 3.5))
    plt.title('null component')
    plt.pie(
        [len(d)-d.isnull().sum(), d.isnull().sum()],
        labels=['value', 'null'],
        colors=['tab:blue', 'tab:orange'],
        startangle=90,
        autopct="%1.3f%%",
        wedgeprops={'linewidth': 1, 'edgecolor':"white"}
    )
    plt.legend(frameon=False, loc='upper left', handlelength=1, handletextpad=0.2, labelspacing=0.2)
    
    ## fig貼り付け
    ee.put_img(ws, fig, f'{d.name}_null', (4,8), img_dir)
    ee.set_style(ws.cell(4,8), 'style_grid')
    ws.merge_cells('H4:K22')
    
    # 幅、高さ調整
    ws.row_dimensions[1].height = 5
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['C'].width = 19
    ws.column_dimensions['E'].width = 19
    
def bar_plot(d):
    fig = plt.figure(figsize=(6,3))
    ax = fig.add_subplot(111)
    ax.set_title(d.name)
    tmp = d.astype(str).fillna('欠損').value_counts().sort_index()
    ax.bar(tmp.index, tmp.values, alpha=0.6)
    ax.set_ylabel('Count')
    ax.set_xlabel('Category')
    plt.xticks(rotation=90)
    ax.set_ylabel('Count')
    ax.set_xlabel('Category')
    ax.grid(linestyle='--', alpha=0.6)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, loc: "{:,}".format(int(x))))
    return fig

def hist_plot(d):
    fig = plt.figure(figsize=(6,3))
    ax = fig.add_subplot(111)
    ax.hist(d, alpha=0.6, bins=50)
    ax.set_ylabel('Frequency')
    ax.set_xlabel('value')
    ax.grid(linestyle='--', alpha=0.6)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, loc: "{:,}".format(int(x))))
    return fig

def too_much_nuniques(d):
    fig = plt.figure(figsize=(6,3))
    ax = fig.add_subplot(111)
    ax.set_title(d.name)
    ax.text(0.5, 0.5, 'nunique > 30', ha='center', va='center')
    ax.grid(linestyle='--', alpha=0.6)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, loc: "{:,}".format(int(x))))
    return fig
    
if __name__ == "__main__":
    '''
    main関数実行
    '''
    main()